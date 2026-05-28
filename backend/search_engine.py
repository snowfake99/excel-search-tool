"""
search_engine.py — in-memory cache + rapidfuzz
"""
import os
import re
from dataclasses import dataclass
from typing import Dict, List, Optional

import openpyxl
from rapidfuzz import fuzz

_CACHE: Dict[str, List[dict]] = {}


def col_index_to_letter(col_index: int) -> str:
    result = ""
    while col_index > 0:
        col_index, remainder = divmod(col_index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def load_file_to_cache(file_path: str) -> int:
    filename = os.path.basename(file_path)
    cells: List[dict] = []
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                original = str(cell.value).strip()
                if not original or original == "nan":
                    continue
                col_letter = col_index_to_letter(cell.column)
                cells.append({
                    "sheet":    str(sheet_name),
                    "row":      int(cell.row),
                    "col":      int(cell.column),
                    "cell_ref": f"{col_letter}{cell.row}",
                    "original": original,
                    "lower":    original.lower(),
                })
    wb.close()
    _CACHE[filename] = cells
    return len(cells)


def remove_file_from_cache(filename: str):
    _CACHE.pop(os.path.basename(filename), None)


def get_cached_files() -> List[str]:
    return list(_CACHE.keys())


def _build_context(cells: List[dict], sheet: str, row: int, col: int) -> str:
    parts = []
    for c in cells:
        if c["sheet"] != sheet:
            continue
        if abs(c["row"] - row) <= 1 and abs(c["col"] - col) <= 1 and (c["row"] != row or c["col"] != col):
            parts.append(f"{c['cell_ref']}:{c['original']}")
        if len(parts) >= 6:
            break
    return " | ".join(parts)


def _compute_score(cell_lower: str, kw_lower: str) -> int:
    if cell_lower == kw_lower:
        return 100
    if kw_lower in cell_lower:
        return 90
    best = max(fuzz.partial_ratio(kw_lower, cell_lower), fuzz.token_sort_ratio(kw_lower, cell_lower))
    return int(best) if best >= 60 else 0


@dataclass
class SearchResult:
    file_name:  str
    sheet_name: str
    row:        int
    col:        int
    cell_ref:   str
    cell_value: str
    context:    str
    score:      int


def search_multiple_files(
    file_paths: List[str],
    keyword: str,
    case_sensitive: bool = False,
    use_regex: bool = False,
    fuzzy: bool = True,
    score_cutoff: int = 60,
    max_workers: int = 4,
) -> List[SearchResult]:
    if not keyword.strip():
        return []

    kw_lower = keyword.lower() if not case_sensitive else keyword

    regex_pattern: Optional[re.Pattern] = None
    if use_regex:
        try:
            flags = 0 if case_sensitive else re.IGNORECASE
            regex_pattern = re.compile(keyword, flags)
        except re.error:
            return []

    all_results: List[SearchResult] = []

    for file_path in file_paths:
        filename = os.path.basename(file_path)
        if filename not in _CACHE:
            try:
                load_file_to_cache(file_path)
            except Exception as e:
                print(f"[ERROR] {filename}: {e}")
                continue

        cells = _CACHE.get(filename, [])

        for c in cells:
            score = 0
            if use_regex and regex_pattern:
                if regex_pattern.search(c["original"]):
                    score = 90
            else:
                cell_cmp = c["lower"] if not case_sensitive else c["original"]
                score = _compute_score(cell_cmp, kw_lower)

            if score >= score_cutoff:
                all_results.append(SearchResult(
                    file_name=  filename,
                    sheet_name= c["sheet"],
                    row=        c["row"],
                    col=        c["col"],
                    cell_ref=   c["cell_ref"],
                    cell_value= c["original"],
                    context=    _build_context(cells, c["sheet"], c["row"], c["col"]),
                    score=      score,
                ))

    all_results.sort(key=lambda x: -x.score)
    return all_results
